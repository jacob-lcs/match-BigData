import xlrd
import openpyxl
from wordcloud import WordCloud
from scipy.misc import imread
import matplotlib.pyplot as plt


def read_excel(file_excel, infos, n, sheetnames):  # 读excel并将需要的数据分类放在数组里
    info_file = xlrd.open_workbook(file_excel)
    info_sheet = info_file.sheets()[n]  # 获取第一个sheet
    row_count = info_sheet.nrows  # 获取数据的行

    for row in range(1, row_count):
        infos.append(
            {
                'position': info_sheet.cell(row, 0).value,  # 获取第一列的值
                'Job': sheetnames[n]
            }
        )
    return infos


if __name__ == "__main__":
    wb = openpyxl.load_workbook('./data/前程无忧_职位.xlsx')
    sheetnames = wb.sheetnames
    for ff in range(8):
        infos = []  # 新建空字典
        print("正在存储第", str(ff + 1), "个sheet...")
        read_excel('./data/八个职位.xlsx', infos, ff, sheetnames)
        print("保存成功！！")
        count = len(infos)  # 总计数
        print("数据总数为：", str(count))
        job = ''
        brief_position = []
        for key in infos:
            brief_position.append(key['position'])
            job = key['Job']

        outstr = ''
        for word in brief_position:  # 去除停用词
            outstr += word
            outstr += "/"

        wc = WordCloud(font_path=r"./fonts/simheittf.ttf",
                       background_color='white',
                       # width=800,
                       # height=600,
                       collocations=False,
                       max_font_size=50,
                       max_words=1000,
                       mask=imread("./data/computer.png"))
        wc.generate(outstr)
        wc.to_file(job + "_职业统计.png")  # 按照设置的像素宽高度保存绘制好的词云图

        # 4、显示图片
        plt.figure(job)  # 指定所绘图名称
        plt.imshow(wc)  # 以图片的形式显示词云
        plt.axis("off")  # 关闭图像坐标系
        plt.show()
