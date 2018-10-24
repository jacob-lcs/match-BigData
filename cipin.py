# -*- coding:utf-8 -*-
import xlrd
import jieba
import xlwt
import openpyxl
from wordcloud import WordCloud  # 词云库
import matplotlib.pyplot as plt  # 数学绘图库


def stopwordslist(filepath):
    stopwords = [line.strip() for line in open(filepath, 'rb', encoding='utf-8').readlines()]
    return stopwords


def excel_write(items, index):
    # 爬取到的内容写入excel表格
    for item in items:  # 职位信息
        ws.write(index, 0, item)
        ws.write(index, 1, items[item])
        index += 1
        if index == 65535:
            break


def read_excel(file_excel, infos, n, sheetnames):  # 读excel并将需要的数据分类放在数组里
    info_file = xlrd.open_workbook(file_excel)
    info_sheet = info_file.sheets()[n]  # 获取第一个sheet
    row_count = info_sheet.nrows  # 获取数据的行

    for row in range(1, row_count):
        infos.append(
            {
                'position': info_sheet.cell(row, 0).value,  # 获取第一列的值
                'company': info_sheet.cell(row, 1).value,
                'addition': info_sheet.cell(row, 2).value,
                'salary': info_sheet.cell(row, 3).value,
                'Education': info_sheet.cell(row, 5).value,
                'work-experience': info_sheet.cell(row, 6).value,
                'Company-Type': info_sheet.cell(row, 7).value,
                'Company-Size': info_sheet.cell(row, 8).value,
                'Job-description': info_sheet.cell(row, 9).value,
                'Job': sheetnames[n]
            }
        )
    return infos


# 创建停用词list
def stopwordslist(filepath):
    stopwords = [line.strip() for line in open(filepath, 'r', encoding='UTF-8').readlines()]
    return stopwords


if __name__ == "__main__":
    quest = [
        ['web前端开发工程师', 'HTML5', 'CSS3', 'JavaScript', 'css', 'css3', 'HTML', 'Ajas', 'ajax', 'ES6', 'es6', 'XML', 'DOM',
         'AngularJS', 'angular', 'vue', 'VUE', 'nodejs', 'NodeJS', 'node', 'Node',
         'ReactJs', 'react', 'ReactJS', 'React', 'element', 'Bootstrap', 'bootstrap', 'Jquery', 'jquery', 'JQuery',
         'jQuery', 'webpack', 'gulp', 'git', '规范', '习惯', '模块化', '前后端分离', '兼容性'],
        ['软件测试'
            , 'JAVA', 'Java'
            , 'Python', 'python'
            , 'selenium', 'Selenium'
            , 'QTP'
            , 'Loadrunner', 'loadrunner'
            , 'jmeter', 'Jmeter'
            , 'mysql', 'Mysql', 'MySQL'
            , 'SQL', 'sql'
            , 'linux', 'Linux'
            , 'APP', 'App'
            , 'web', 'WEB'
            , 'Android', 'android'
            , '数据库'
            , '文档编写'
         ],
        ['ERP技术开发', 'ERP', 'oracle', 'ORACLE', 'SQL', 'SAP', 'ABAP', 'JAVA', 'Java', 'java', 'NET', 'net', 'Net',
         'Delphi', 'C#', 'Delphi', 'CRM', '金蝶', '用友', '二次开发', '沟通', '前端'],
        ['语音视频图形开发', 'C', 'C++', 'c++', '音视频开发', '语音识别', 'kaidi', 'Kaldi', '图像处理', '计算机视觉', '编解码', '深度学习', 'TensorFlow',
         'RTMP', 'rtmp', 'RTSP', 'rtsp', 'WebRTC', 'webrtc', 'ffmpeg', 'FFMPEG', 'H264', 'h264', 'h.264', 'aac', 'AAC',
         'Android', 'iOS', 'OpenCV'],
        ['系统工程师',
         'HP', 'hp',
         'IBM', 'ibm',
         'ERP',
         'Spring', 'spring',
         'J2EE',
         'Linux', 'linux',
         'MYSQL', 'MySQL', 'mysql', 'Mysql',
         'shell', 'Shell',
         'python', 'Python',
         'SAP',
         'DNS',
         'JAVA', 'java',
         'c++', 'C++',
         'OA',
         'Unix', 'unix', 'UNIX',
         'Socket', 'socket',
         'MCSE',
         'RHCE',
         'Nginx', 'nginx',
         'Tomcat', 'tomcat',
         'Ansible', 'ansible',
         'Spring', 'spring', 'SpringMVC', 'springmvc',
         'MES',
         'Redis', 'redis',
         'AWS',
         'VMWARE', 'VMwaremware',
         'Docker',
         'Spark',
         'PLC',
         'Hadoop',
         'Zabbix', 'zabbix',
         '华为',
         '分布式',
         '交换机',
         '虚拟化',
         '路由器', '路由',
         '防火墙',
         '备份',
         '排查',
         '英文阅读能力'
         ],
        ['项目总监',
         '制定',
         '计划',
         '项目',
         '成本',
         '风险',
         '需求',
         '调研', '调查',
         '经验',
         '沟通',
         '逻辑思维',
         '压力',
         'PMP',
         'J2EE',
         'CMM3',
         '整体把控',
         '判断',
         'WBS'
         ],
        ['需求工程师',
         '分析',
         '调研',
         '产品设计',
         '交互', 'UI',
         '大型', '大型项目',
         '文档撰写',
         '数据库',
         '团队', '团队精神',
         '逻辑思维', '逻辑',
         '用户',
         '业务流程',
         'Axure', 'axure',
         'visio', 'Visio', 'VISIO',
         'uml', 'UML',
         'SQL',
         '英语', '英文',
         '创新',
         '沟通'
         ],
        ['数据库工程师'
            , 'Linux', 'linux'
            , 'SQL', 'sql'
            , 'ELT'
            , 'Hadoop', 'hadoop'
            , 'Scala', 'scala'
            , 'Redis', 'redis'
            , '设计'
            , '平台'
            , '数据仓库'
            , '逻辑'
            , '沟通'
            , '抗压能力'
            , '突发事件'
         ]]
    wb = openpyxl.load_workbook('./data/前程无忧_职位.xlsx')
    sheetnames = wb.sheetnames
    list = {'web前端开发工程师',
            '软件测试',
            'ERP技术开发',
            '语音视频图形开发',
            '系统工程师',
            '项目总监',
            '需求工程师',
            '数据库工程师'}
    for ff in range(57):
        if sheetnames[ff] in list:
            infos = []  # 新建空字典
            print("正在存储第", str(ff + 1), "个sheet...", "sheet表名为:", sheetnames[ff])
            read_excel('./data/前程无忧_职位.xlsx', infos, ff, sheetnames)
            print("保存成功！！")
            count = len(infos)  # 总计数
            print("数据总数为：", str(count))
            job = ''
            brief_position = []
            for key in infos:
                brief_position.append(key['Job-description'])
                job = key['Job']

            word_count = {}
            for words in quest:
                if sheetnames[ff] == words[0]:
                    print(words[0])
                    for key in brief_position:
                        for word in words:
                            if word in key:
                                if word in word_count:
                                    word_count[word] += 1
                                else:
                                    word_count[word] = 1

            newTable = job + ".xls"  # 表格名称
            wb = xlwt.Workbook(encoding='utf-8')  # 创建excel文件，声明编码
            ws = wb.add_sheet(job)  # 创建表格
            headData = ['要求', '计数', str(count)]  # 表头部信息
            for colnum in range(0, 3):
                ws.write(0, colnum, headData[colnum], xlwt.easyxf('font: bold on'))  # 行，列
            length = len(word_count)
            excel_write(word_count, 1)
            wb.save(newTable)
