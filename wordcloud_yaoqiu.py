from wordcloud import WordCloud
import matplotlib.pyplot as plt
import xlrd
import openpyxl
from scipy.misc import imread


def read_excel(file_excel, infos, n, sheetnames):  # 读excel并将需要的数据分类放在数组里
    info_file = xlrd.open_workbook(file_excel)
    info_sheet = info_file.sheets()[n]  # 获取第一个sheet
    row_count = info_sheet.nrows  # 获取数据的行

    for row in range(1, row_count):
        infos.append(
            {
                'keyword': info_sheet.cell(row, 0).value,  # 获取第一列的值
                'amount': info_sheet.cell(row, 1).value,
                'Job': sheetnames[n]
            }
        )
    return infos


if __name__ == "__main__":
    wb = openpyxl.load_workbook('./data/词频统计.xlsx')
    sheetnames = wb.sheetnames
    for ff in range(8):
        infos = []  # 新建空字典
        print("正在存储第", str(ff + 1), "个sheet...")
        read_excel('./data/词频统计.xlsx', infos, ff, sheetnames)
        print("保存成功！！")
        count = len(infos)  # 总计数
        print("数据总数为：", str(count))
        job = ''
        brief_position = []
        amount = []
        for key in infos:
            brief_position.append(key['keyword'])
            amount.append(key['amount'])
            job = key['Job']
        i = 0
        outstr = ''
        for k in brief_position:
            for i in range(int(amount[1])):
                outstr += k
                outstr += "/"
            i += 1

        wc = WordCloud(font_path=r"./fonts/simheittf.ttf",
                       background_color='white',
                       # width=800,
                       # height=600,
                       collocations=False,
                       max_font_size=50,
                       max_words=1000,
                       mask=imread("./data/computer.png"))
        wc.generate(outstr)
        wc.to_file(job + ".png")  # 按照设置的像素宽高度保存绘制好的词云图

        # 4、显示图片
        plt.figure(job)  # 指定所绘图名称
        plt.imshow(wc)  # 以图片的形式显示词云
        plt.axis("off")  # 关闭图像坐标系
        plt.show()
