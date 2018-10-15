import xlrd
import re
import xlwt
import openpyxl


def read_excel(file_excel, infos, n, SheetNames):  # 读excel并将需要的数据分类放在数组里
    info_file = xlrd.open_workbook(file_excel)
    info_sheet = info_file.sheets()[n]  # 获取sheet
    row_count = info_sheet.nrows  # 获取数据的行
    for row in range(1, row_count):
        s = info_sheet.cell(row, 3).value
        # print(s)
        # ----------------工资---------------------------#
        if "万/月" in s:
            matchObj = re.match(r'(.*)-(.*)万/月', s, re.M | re.I)
            m1 = matchObj.group(1)
            m2 = matchObj.group(2)
            aver = (float(m1) + float(m2)) / 2 * 10000
        elif "千/月" in s:
            matchObj = re.match(r'(.*)-(.*)千/月', s, re.M | re.I)
            m1 = matchObj.group(1)
            m2 = matchObj.group(2)
            aver = (float(m1) + float(m2)) / 2 * 1000
        elif "元/天" in s:
            matchObj = re.match(r'(.*)元/天', s, re.M | re.I)
            m1 = matchObj.group(1)
            aver = float(m1) * 30
        elif s == '':
            aver = 5000
        elif "千以下/月" in s:
            matchObj = re.match(r'(.*)千以下/月', s, re.M | re.I)
            m1 = matchObj.group(1)
            aver = float(m1)
        elif "万/年" in s:
            matchObj = re.match(r'(.*)-(.*)万/年', s, re.M | re.I)
            m1 = matchObj.group(1)
            m2 = matchObj.group(2)
            aver = (float(m1) + float(m2)) / 2 * 10000 / 12
        # ----------------工资---------------------------#
        infos.append(
            {
                'Big_Position': SheetNames[n],
                'position': info_sheet.cell(row, 0).value,  # 获取第一列的值
                'company': info_sheet.cell(row, 1).value,
                'addition': info_sheet.cell(row, 2).value,
                'salary': aver,
                'Education': info_sheet.cell(row, 5).value,
                'work-experience': info_sheet.cell(row, 6).value,
                'Company-Type': info_sheet.cell(row, 7).value,
                'Company-Size': info_sheet.cell(row, 8).value,
                'Job-description': info_sheet.cell(row, 9).value
            }
        )
    return infos


def excel_write(items):
    # 爬取到的内容写入excel表格
    len = 1
    for item in items:  # 职位信息
        ws.write(len, 0, item['addition'])  # 行，列，数据
        ws.write(len, 1, item['技术文员'])
        ws.write(len, 2, item['计算机辅助设计工程师'])
        ws.write(len, 3, item['高级软件工程师'])
        ws.write(len, 4, item['ERP技术开发'])
        ws.write(len, 5, item['ERP实施顾问'])
        ws.write(len, 6, item['仿真应用工程师'])
        ws.write(len, 7, item['软件UI设计师'])
        ws.write(len, 8, item['软件工程师'])
        ws.write(len, 9, item['算法工程师'])
        ws.write(len, 10, item['数据库工程师'])
        ws.write(len, 11, item['系统分析员'])
        ws.write(len, 12, item['系统集成工程师'])
        ws.write(len, 13, item['系统工程师'])
        ws.write(len, 14, item['需求工程师'])
        ws.write(len, 15, item['系统架构设计师'])
        ws.write(len, 16, item['网站维护工程师'])
        ws.write(len, 17, item['游戏开发工程师'])
        ws.write(len, 18, item['系统网络管理员'])
        ws.write(len, 19, item['硬件工程师'])
        ws.write(len, 20, item['手机应用开发工程师'])
        ws.write(len, 21, item['网站架构设计师'])
        ws.write(len, 22, item['脚本开发工程师'])
        ws.write(len, 23, item['网络信息安全工程师'])
        ws.write(len, 24, item['网络工程师'])
        ws.write(len, 25, item['高级硬件工程师'])
        ws.write(len, 26, item['大数据开发工程师'])
        ws.write(len, 27, item['互联网软件开发工程师'])
        ws.write(len, 28, item['web前端开发工程师'])
        ws.write(len, 29, item['语音视频图形开发'])
        ws.write(len, 30, item['计量工程师'])
        ws.write(len, 31, item['电脑维修'])
        ws.write(len, 32, item['标准化工程师'])
        ws.write(len, 33, item['测试员'])
        ws.write(len, 34, item['技术支持维护工程师'])
        ws.write(len, 35, item['技术总监'])
        ws.write(len, 36, item['技术支持维护经理'])
        ws.write(len, 37, item['视觉设计师'])
        ws.write(len, 38, item['配置管理工程师'])
        ws.write(len, 39, item['手机维修'])
        ws.write(len, 40, item['软件测试'])
        ws.write(len, 41, item['网络管理'])
        ws.write(len, 42, item['首席技术执行官'])
        ws.write(len, 43, item['特效设计师'])
        ws.write(len, 44, item['网络维修'])
        ws.write(len, 45, item['系统测试'])
        ws.write(len, 46, item['项目主管'])
        ws.write(len, 47, item['项目经理'])
        ws.write(len, 48, item['网页设计'])
        ws.write(len, 49, item['项目执行、协调人员'])
        ws.write(len, 50, item['信息技术经理'])
        ws.write(len, 51, item['信息技术专员'])
        ws.write(len, 52, item['项目总监'])
        ws.write(len, 53, item['音效设计师'])
        ws.write(len, 54, item['用户体验设计师'])
        ws.write(len, 55, item['硬件测试'])
        ws.write(len, 56, item['UI设计师'])
        ws.write(len, 57, item['Flash设计'])
        len += 1


if __name__ == "__main__":
    infos = []  # 新建空字典
    wb = openpyxl.load_workbook('./data/前程无忧_职位.xlsx')
    SheetNames = wb.sheetnames
    for n in range(57):  # 存储sheet
        print("正在存储第", str(n + 1), "个sheet...")
        read_excel('./data/前程无忧_职位.xlsx', infos, n, SheetNames)
    print("保存成功！！")
    count = len(infos)  # 总计数
    print("数据总数为：", str(count))

    cs2 = []  # 新建空字典
    total = 0
    for key in infos:
        addition = key['addition']
        salary = key['salary']
        position = key['Big_Position']
        confirm = 0
        for lcs in cs2:
            if lcs['addition'] == addition:  # 遍历到的地址在字典中
                confirm = 1
                break
        if confirm == 0:
            c = 0
            if '-' in addition:
                matchObj = re.match(r'(.*)-.*', addition, re.M | re.I)  # 正则表达式匹配城市名
                if matchObj:
                    addition = matchObj.group(1)
                c = 0
            for lcs in cs2:
                if lcs['addition'] == addition:  # 遍历到的地址在字典中
                    c = 1
                    break
            if c == 0:
                addition2 = {'addition': addition, '技术文员': 0, '计算机辅助设计工程师': 0, '高级软件工程师': 0, 'ERP技术开发': 0, 'ERP实施顾问': 0,
                             '仿真应用工程师': 0, '软件UI设计师': 0, '软件工程师': 0, '算法工程师': 0, '数据库工程师': 0, '系统分析员': 0, '系统集成工程师': 0,
                             '系统工程师': 0, '需求工程师': 0, '系统架构设计师': 0, '网站维护工程师': 0, '游戏开发工程师': 0, '系统网络管理员': 0, '硬件工程师': 0,
                             '手机应用开发工程师': 0, '网站架构设计师': 0, '脚本开发工程师': 0, '网络信息安全工程师': 0, '网络工程师': 0, '高级硬件工程师': 0,
                             '大数据开发工程师': 0, '互联网软件开发工程师': 0, 'web前端开发工程师': 0, '语音视频图形开发': 0, '计量工程师': 0, '电脑维修': 0,
                             '标准化工程师': 0, '测试员': 0, '技术支持维护工程师': 0, '技术总监': 0, '技术支持维护经理': 0, '视觉设计师': 0, '配置管理工程师': 0,
                             '手机维修': 0, '软件测试': 0, '网络管理': 0, '首席技术执行官': 0, '特效设计师': 0, '网络维修': 0, '系统测试': 0, '项目主管': 0,
                             '项目经理': 0, '网页设计': 0, '项目执行、协调人员': 0, '信息技术经理': 0, '信息技术专员': 0, '项目总监': 0, '音效设计师': 0,
                             '用户体验设计师': 0, '硬件测试': 0, 'UI设计师': 0, 'Flash设计': 0, '技术文员1': 0, '计算机辅助设计工程师1': 0, '高级软件工程师1': 0,
                             'ERP技术开发1': 0, 'ERP实施顾问1': 0,
                             '仿真应用工程师1': 0, '软件UI设计师1': 0, '软件工程师1': 0, '算法工程师1': 0, '数据库工程师1': 0, '系统分析员1': 0, '系统集成工程师1': 0,
                             '系统工程师1': 0, '需求工程师1': 0, '系统架构设计师1': 0, '网站维护工程师1': 0, '游戏开发工程师1': 0, '系统网络管理员1': 0, '硬件工程师1': 0,
                             '手机应用开发工程师1': 0, '网站架构设计师1': 0, '脚本开发工程师1': 0, '网络信息安全工程师1': 0, '网络工程师1': 0, '高级硬件工程师1': 0,
                             '大数据开发工程师1': 0, '互联网软件开发工程师1': 0, 'web前端开发工程师1': 0, '语音视频图形开发1': 0, '计量工程师1': 0, '电脑维修1': 0,
                             '标准化工程师1': 0, '测试员1': 0, '技术支持维护工程师1': 0, '技术总监1': 0, '技术支持维护经理1': 0, '视觉设计师1': 0, '配置管理工程师1': 0,
                             '手机维修1': 0, '软件测试1': 0, '网络管理1': 0, '首席技术执行官1': 0, '特效设计师1': 0, '网络维修1': 0, '系统测试1': 0, '项目主管1': 0,
                             '项目经理1': 0, '网页设计1': 0, '项目执行、协调人员1': 0, '信息技术经理1': 0, '信息技术专员1': 0, '项目总监1': 0, '音效设计师1': 0,
                             '用户体验设计师1 ': 0, '硬件测试1': 0, 'UI设计师1': 0, 'Flash设计1': 0}
                cs2.append(addition2)
        for ad in cs2:
            if ad["addition"] == addition:
                ad[position] += salary
                p = position + "1"
                ad[p] += 1
    # print(cs2)
    for ad in cs2:
        for a in ad:
            if a != "addition":
                if "1" not in a:
                    if ad[a] != 0:
                        aa = a + "1"
                        ad[a] = ad[a] / ad[aa]
    newTable = "各城市平均工资统计.xls"  # 表格名称
    wb = xlwt.Workbook(encoding='utf-8')  # 创建excel文件，声明编码
    ws = wb.add_sheet('各城市平均工资统计')  # 创建表格
    headData = ['addition', '技术文员', '计算机辅助设计工程师', '高级软件工程师', 'ERP技术开发', 'ERP实施顾问',
                '仿真应用工程师', '软件UI设计师', '软件工程师', '算法工程师', '数据库工程师', '系统分析员', '系统集成工程师',
                '系统工程师', '需求工程师', '系统架构设计师', '网站维护工程师', '游戏开发工程师', '系统网络管理员', '硬件工程师',
                '手机应用开发工程师', '网站架构设计师', '脚本开发工程师', '网络信息安全工程师', '网络工程师', '高级硬件工程师',
                '大数据开发工程师', '互联网软件开发工程师', 'web前端开发工程师', '语音视频图形开发', '计量工程师', '电脑维修',
                '标准化工程师', '测试员', '技术支持维护工程师', '技术总监', '技术支持维护经理', '视觉设计师', '配置管理工程师',
                '手机维修', '软件测试', '网络管理', '首席技术执行官', '特效设计师', '网络维修', '系统测试', '项目主管',
                '项目经理', '网页设计', '项目执行、协调人员', '信息技术经理', '信息技术专员', '项目总监', '音效设计师',
                '用户体验设计师', '硬件测试', 'UI设计师', 'Flash设计']  # 表头部信息
    for colnum in range(0, 58):
        ws.write(0, colnum, headData[colnum], xlwt.easyxf('font: bold on'))  # 行，列
    excel_write(cs2)
    wb.save(newTable)
