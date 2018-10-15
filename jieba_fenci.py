# -*- coding:utf-8 -*-
import xlrd
import jieba
import xlwt
import openpyxl
from wordcloud import WordCloud  # 词云库
import matplotlib.pyplot as plt  # 数学绘图库


def stopwordslist(filepath):
    stopwords = [line.strip() for line in open(filepath, 'rb', encoding='utf-8').readlines()]
    return stopwords


def excel_write(items, index):
    # 爬取到的内容写入excel表格
    for item in items:  # 职位信息
        ws.write(index, 0, item)
        ws.write(index, 1, items[item])
        index += 1
        if index == 65535:
            break


def read_excel(file_excel, infos, n, sheetnames):  # 读excel并将需要的数据分类放在数组里
    info_file = xlrd.open_workbook(file_excel)
    info_sheet = info_file.sheets()[n]  # 获取第一个sheet
    row_count = info_sheet.nrows  # 获取数据的行

    for row in range(1, row_count):
        infos.append(
            {
                'position': info_sheet.cell(row, 0).value,  # 获取第一列的值
                'company': info_sheet.cell(row, 1).value,
                'addition': info_sheet.cell(row, 2).value,
                'salary': info_sheet.cell(row, 3).value,
                'Education': info_sheet.cell(row, 5).value,
                'work-experience': info_sheet.cell(row, 6).value,
                'Company-Type': info_sheet.cell(row, 7).value,
                'Company-Size': info_sheet.cell(row, 8).value,
                'Job-description': info_sheet.cell(row, 9).value,
                'Job': sheetnames[n]
            }
        )
    return infos


# 创建停用词list
def stopwordslist(filepath):
    stopwords = [line.strip() for line in open(filepath, 'r', encoding='UTF-8').readlines()]
    return stopwords


if __name__ == "__main__":
    wb = openpyxl.load_workbook('./data/前程无忧_职位.xlsx')
    sheetnames = wb.sheetnames
    for ff in range(57):
        infos = []  # 新建空字典
        print("正在存储第", str(ff + 1), "个sheet...")
        read_excel('./data/前程无忧_职位.xlsx', infos, ff, sheetnames)
        print("保存成功！！")
        count = len(infos)  # 总计数
        print("数据总数为：", str(count))
        job = ''
        brief_position = []
        for key in infos:
            brief_position.append(key['Job-description'])
            job = key['Job']

        brief_position_str = ''.join(brief_position)
        wordlist = jieba.cut(brief_position_str, cut_all=False)
        stopwords = stopwordslist('./stopwords/哈工大停用词表.txt')  # 这里加载停用词的路径
        outstr = ''
        word_count = {}
        for word in wordlist:  # 去除停用词
            if word not in stopwords:
                if word != '\t':
                    outstr += word
                    outstr += "/"
                    if word in word_count:
                        word_count[word] += 1
                    else:
                        word_count[word] = 1

        newTable = job + ".xls"  # 表格名称
        wb = xlwt.Workbook(encoding='utf-8')  # 创建excel文件，声明编码
        ws = wb.add_sheet(job)  # 创建表格
        headData = ['要求', '计数']  # 表头部信息
        for colnum in range(0, 2):
            ws.write(0, colnum, headData[colnum], xlwt.easyxf('font: bold on'))  # 行，列
        length = len(word_count)
        excel_write(word_count, 1)
        wb.save(newTable)

# #################################  云图  ##################################
# wc = WordCloud(font_path=r"./fonts/simheittf.ttf", background_color='white', width=800,
#                height=600, max_font_size=50,
#                max_words=1000)
# wc.generate(outstr)
# wc.to_file(r"wordcloud.png")  # 按照设置的像素宽高度保存绘制好的词云图
#
# # 4、显示图片
# plt.figure("词云图")  # 指定所绘图名称
# plt.imshow(wc)  # 以图片的形式显示词云
# plt.axis("off")  # 关闭图像坐标系
# plt.show()

# #################################  云图  ##################################
